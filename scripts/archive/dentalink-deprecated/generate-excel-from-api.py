#!/usr/bin/env python3
"""
Genera Excel de pagos desde API de HealthAtom (token permanente).
NO necesita PHPSESSID ni sesión web.

Primero intenta fetch fresco desde API. Si falla por rate-limit,
usa los datos existentes en public/pagos-data.json.

Usage:
  DENTALINK_TOKEN=xxx python3 scripts/generate-excel-from-api.py
  python3 scripts/generate-excel-from-api.py --use-existing  # solo datos guardados
"""
import os, sys, json, urllib.request, urllib.error, time
from datetime import datetime

BASE = "https://api.dentalink.healthatom.com/api/v1"
TOKEN = os.environ.get("DENTALINK_TOKEN") or os.environ.get("DENTALINK_API_TOKEN")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(ROOT, "data", "imports", "pagos")
os.makedirs(OUT_DIR, exist_ok=True)

# Load from .env.dentalink-web if not set
if not TOKEN:
    env_file = os.path.join(ROOT, ".env.dentalink-web")
    if os.path.exists(env_file):
        with open(env_file) as f:
            for line in f:
                line = line.strip()
                if line.startswith("DENTALINK_TOKEN="):
                    TOKEN = line.split("=", 1)[1].strip().strip('"').strip("'")
                    break

if not TOKEN:
    print("❌ DENTALINK_TOKEN no encontrado. Pásalo como env var o ponlo en .env.dentalink-web")
    sys.exit(1)

USE_EXISTING = "--use-existing" in sys.argv

def fetch_all(endpoint, delay=1500, max_retries=8):
    """Paginación completa con retry en rate-limit."""
    results = []
    url = f"{BASE}/{endpoint}"
    page = 0
    while url:
        page += 1
        ok = False
        for attempt in range(max_retries):
            try:
                req = urllib.request.Request(url)
                req.add_header("Authorization", f"Token {TOKEN}")
                req.add_header("Accept", "application/json")
                resp = urllib.request.urlopen(req, timeout=30)
                data = json.loads(resp.read().decode())
                ok = True
                break
            except urllib.error.HTTPError as e:
                if e.code == 429 and attempt < max_retries - 1:
                    wait = min(30, 5 * (attempt + 1))
                    print(f"\n  ⚠️ Rate limit page {page}, esperando {wait}s (intento {attempt+2}/{max_retries})")
                    time.sleep(wait)
                    continue
                else:
                    print(f"  ⚠️ Error page {page}: {e}")
                    return results
            except Exception as e:
                print(f"  ⚠️ Error page {page}: {e}")
                return results
        
        if not ok:
            break
            
        batch = data.get("data", [])
        results.extend(batch)
        next_url = None
        if isinstance(data.get("links"), dict):
            next_url = data["links"].get("next")
        url = next_url
        
        if len(results) % 100 == 0 or len(results) < 50:
            print(f"  Page {page}: {len(results)} total", end="\r")
        if url:
            time.sleep(delay / 1000)
    
    print(f"\n  ✅ {len(results)} registros obtenidos")
    return results


def get_tratamientos(pids, delay=60):
    """Obtiene tratamientos/acciones de pacientes."""
    acciones = {}
    total = len(pids)
    for i, pid in enumerate(pids):
        url = f"{BASE}/pacientes/{pid}/tratamientos"
        for attempt in range(3):
            try:
                req = urllib.request.Request(url)
                req.add_header("Authorization", f"Token {TOKEN}")
                req.add_header("Accept", "application/json")
                resp = urllib.request.urlopen(req, timeout=10)
                data = json.loads(resp.read().decode())
                items = data.get("data", [])
                if items:
                    acciones[pid] = items
                if (i+1) % 100 == 0 or i == total - 1:
                    print(f"  {i+1}/{total}\r", end="")
                time.sleep(delay / 1000)
                break
            except Exception:
                time.sleep(1)
                continue
    print(f"\n  ✅ Acciones de {len(acciones)} pacientes")
    return acciones


def make_excel(pagos, filename, columns, transform_fn=None):
    """Genera Excel con estilo formateado."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        os.system("pip install openpyxl -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws.freeze_panes = "A2"
    
    for row_idx, pago in enumerate(pagos, 2):
        if transform_fn:
            row_data = transform_fn(pago)
        else:
            row_data = [pago.get(c, "") for c in columns]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
    
    # Auto-width columns
    for col_idx in range(1, len(columns) + 1):
        max_len = len(str(columns[col_idx - 1]))
        for row_idx in range(2, min(len(pagos) + 2, 50)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 3
    
    path = os.path.join(OUT_DIR, filename)
    wb.save(path)
    print(f"  ✅ Guardado: {path}")


# ========== MAIN ==========

pagos = []

if USE_EXISTING:
    # Skip API, use existing data
    existing = os.path.join(ROOT, "public", "pagos-data.json")
    if os.path.exists(existing):
        print("📂 Usando datos existentes (--use-existing)...")
        with open(existing) as f:
            data = json.load(f)
        pagos = data.get("pagos", [])
        print(f"   {len(pagos)} pagos cargados")
else:
    # Try API first
    print("1️⃣  Obteniendo pagos desde API...")
    pagos = fetch_all("pagos")
    
    # Fallback if API returned empty
    if not pagos:
        existing = os.path.join(ROOT, "public", "pagos-data.json")
        if os.path.exists(existing):
            print(f"\n⚠️  Usando datos existentes de {existing}")
            with open(existing) as f:
                data = json.load(f)
            pagos = data.get("pagos", [])
            print(f"   {len(pagos)} pagos cargados")

if not pagos:
    print("❌ No se pudieron obtener pagos")
    sys.exit(1)

print(f"\n📈 {len(pagos)} pagos obtenidos")

# Calculate date range
dates = [p.get("fecha_recepcion", "") for p in pagos if p.get("fecha_recepcion")]
dates = [d for d in dates if d]
date_min = min(dates) if dates else "?"
date_max = max(dates) if dates else "?"

# === 2. PAGOS GLOBALES ===
print("\n2️⃣  Generando Excel PAGOS GLOBALES...")
global_columns = [
    "ID Pago", "ID Paciente", "Paciente", "Monto",
    "Medio Pago", "Fecha Recepción", "Fecha Creación",
    "Referencia", "Sucursal", "Folio"
]

def global_transform(p):
    return [
        p.get("id", ""),
        p.get("id_paciente", ""),
        p.get("nombre_paciente", ""),
        float(p.get("monto_pago", 0)),
        p.get("medio_pago", ""),
        p.get("fecha_recepcion", ""),
        p.get("fecha_creacion", ""),
        p.get("numero_referencia", ""),
        p.get("nombre_sucursal", ""),
        p.get("folio", "")
    ]

make_excel(pagos, "pagos-globales-api.xlsx", global_columns, global_transform)

# === 3. PAGOS DETALLE ACCIÓN ===
print("\n3️⃣  Generando Excel PAGOS DETALLE ACCIÓN...")

pids = list(set(int(p["id_paciente"]) for p in pagos if p.get("id_paciente")))
print(f"\nObteniendo acciones de {len(pids)} pacientes...")
acciones = get_tratamientos(pids, delay=60)

# Build detail rows
detalle_rows = []
for p in pagos:
    pid = int(p.get("id_paciente", 0))
    acciones_paciente = acciones.get(pid, [{"nombre": "", "monto": 0, "fecha": ""}])
    
    for acc in acciones_paciente:
        detalle_rows.append({
            "id_pago": p.get("id", ""),
            "id_paciente": pid,
            "paciente": p.get("nombre_paciente", ""),
            "monto_pago": float(p.get("monto_pago", 0)),
            "medio_pago": p.get("medio_pago", ""),
            "fecha_recepcion": p.get("fecha_recepcion", ""),
            "servicio": acc.get("nombre", acc.get("accion", "")) if isinstance(acc, dict) else str(acc),
            "monto_servicio": float(acc.get("monto", 0)) if isinstance(acc, dict) else 0,
            "fecha_servicio": acc.get("fecha", "") if isinstance(acc, dict) else "",
            "sucursal": p.get("nombre_sucursal", ""),
            "referencia": p.get("numero_referencia", "")
        })

detalle_columns = [
    "ID Pago", "ID Paciente", "Paciente", "Monto Pago",
    "Medio Pago", "Fecha Pago", "Servicio", "Monto Servicio",
    "Fecha Servicio", "Sucursal", "Referencia"
]

def detalle_transform(d):
    return [
        d["id_pago"], d["id_paciente"], d["paciente"], d["monto_pago"],
        d["medio_pago"], d["fecha_recepcion"], d["servicio"],
        d["monto_servicio"], d["fecha_servicio"], d["sucursal"], d["referencia"]
    ]

make_excel(detalle_rows, "pagos-detalle-api.xlsx", detalle_columns, detalle_transform)

print(f"\n{'='*60}")
print("✅ EXCEL GENERADOS EXITOSAMENTE!")
print(f"{'='*60}")
print(f"  Rango: {date_min} → {date_max}")
print(f"  Total pagos: {len(pagos)}")
print(f"  Total filas detalle: {len(detalle_rows)}")
