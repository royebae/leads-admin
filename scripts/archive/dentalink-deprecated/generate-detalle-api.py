#!/usr/bin/env python3
"""Genera el Excel de Detalle por Acción usando API en paralelo."""
import os, sys, json, urllib.request, time
from concurrent.futures import ThreadPoolExecutor, as_completed

TOKEN = os.environ.get("DENTALINK_TOKEN") or ""
if not TOKEN:
    env_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env.dentalink-web")
    if os.path.exists(env_file):
        with open(env_file) as f:
            for line in f:
                if line.startswith("DENTALINK_TOKEN="):
                    TOKEN = line.split("=", 1)[1].strip().strip('"\'')

BASE = "https://api.dentalink.healthatom.com/api/v1"
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(ROOT, "data", "imports", "pagos")
os.makedirs(OUT_DIR, exist_ok=True)

def fetch_url(url, timeout=10):
    req = urllib.request.Request(url)
    req.add_header("Authorization", f"Token {TOKEN}")
    req.add_header("Accept", "application/json")
    try:
        resp = urllib.request.urlopen(req, timeout=timeout)
        return json.loads(resp.read().decode())
    except Exception:
        return None

def get_all_tratamientos(pids, max_workers=20):
    """Fetch tratamientos for all patients in parallel."""
    results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(fetch_url, f"{BASE}/pacientes/{pid}/tratamientos?limit=100"): pid for pid in pids}
        done = 0
        for f in as_completed(futures):
            pid = futures[f]
            done += 1
            try:
                data = f.result()
                if data and data.get("data"):
                    results[pid] = data["data"]
            except Exception:
                pass
            if done % 200 == 0 or done == len(pids):
                print(f"  {done}/{len(pids)} pacientes", end="\r")
    print(f"\n  ✅ Tratamientos de {len(results)}/{len(pids)} pacientes")
    return results

# Load pagos
with open(os.path.join(ROOT, "public", "pagos-data.json")) as f:
    data = json.load(f)
pagos = data["pagos"]
print(f"📊 {len(pagos)} pagos cargados")

pids = list(set(int(p["id_paciente"]) for p in pagos if p.get("id_paciente")))
print(f"👥 {len(pids)} pacientes únicos")

tratamientos = get_all_tratamientos(pids)

# Build detalle rows
detalle_rows = []
for p in pagos:
    pid = int(p.get("id_paciente", 0))
    trata_list = tratamientos.get(pid, [])
    
    if trata_list:
        for t in trata_list:
            detalle_rows.append({
                "id_pago": p.get("id", ""),
                "id_paciente": pid,
                "paciente": p.get("nombre_paciente", ""),
                "monto_pago": float(p.get("monto_pago", 0)),
                "medio_pago": p.get("medio_pago", ""),
                "fecha_pago": p.get("fecha_recepcion", ""),
                "servicio": t.get("nombre", ""),
                "monto_servicio": float(t.get("total", 0)),
                "fecha_servicio": t.get("fecha", ""),
                "sucursal": p.get("nombre_sucursal", ""),
                "referencia": p.get("numero_referencia", ""),
                "dentista": t.get("nombre_dentista", ""),
                "finalizado": "Sí" if t.get("finalizado") else "No"
            })
    else:
        detalle_rows.append({
            "id_pago": p.get("id", ""),
            "id_paciente": pid,
            "paciente": p.get("nombre_paciente", ""),
            "monto_pago": float(p.get("monto_pago", 0)),
            "medio_pago": p.get("medio_pago", ""),
            "fecha_pago": p.get("fecha_recepcion", ""),
            "servicio": "",
            "monto_servicio": 0,
            "fecha_servicio": "",
            "sucursal": p.get("nombre_sucursal", ""),
            "referencia": p.get("numero_referencia", ""),
            "dentista": "",
            "finalizado": ""
        })

print(f"📝 {len(detalle_rows)} filas de detalle")

# Generate Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

columns = ["ID Pago", "ID Paciente", "Paciente", "Monto Pago", "Medio Pago",
           "Fecha Pago", "Servicio", "Monto Servicio", "Fecha Servicio",
           "Sucursal", "Referencia", "Dentista", "Finalizado"]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='CBD5E0'),
    right=Side(style='thin', color='CBD5E0'),
    top=Side(style='thin', color='CBD5E0'),
    bottom=Side(style='thin', color='CBD5E0')
)

wb = Workbook()
ws = wb.active
ws.title = "Detalle Acción"

for ci, col in enumerate(columns, 1):
    c = ws.cell(1, ci, col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border

for ri, d in enumerate(detalle_rows, 2):
    vals = [d["id_pago"], d["id_paciente"], d["paciente"], d["monto_pago"],
            d["medio_pago"], d["fecha_pago"], d["servicio"], d["monto_servicio"],
            d["fecha_servicio"], d["sucursal"], d["referencia"], d["dentista"], d["finalizado"]]
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(ri, ci, v)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        if isinstance(v, (int, float)):
            cell.number_format = '#,##0.00'

ws.freeze_panes = "A2"
path = os.path.join(OUT_DIR, "pagos-detalle-api.xlsx")
wb.save(path)
print(f"✅ {path} ({len(detalle_rows)} filas, {len(pids)} pacientes)")
